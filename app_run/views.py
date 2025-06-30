from django.conf import settings
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import viewsets
from openpyxl import load_workbook
from .serializers import RunSerializer, UserSerializer, AthleteInfoSerializer, \
    AthleteDetailSerializer
from .serializers import ChallengeSerializer, PositionSerializer, \
    CollectibleItemSerializer
from .models import Run, AthleteInfo, Challenge, Position, CollectibleItem
from django.contrib.auth.models import User
from project_run.settings import base
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.views import APIView
from rest_framework import status
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination
from geopy.distance import geodesic
from django.db.models import Sum, Count, Q


@api_view(['GET'])
def company_details(request):
    return Response({
        'company_name': base.COMPANY_NAME,
        'slogan': base.SLOGAN,
        'contacts': base.CONTACTS
    })


class RunPagination(PageNumberPagination):
    page_size_query_param = 'size'


class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.select_related('athlete').all()
    serializer_class = RunSerializer
    pagination_class = RunPagination
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'athlete']
    ordering_fields = ['created_at']


class UserPagination(PageNumberPagination):
    page_size_query_param = 'size'


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    pagination_class = UserPagination
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['first_name', 'last_name']
    ordering_fields = ['date_joined']

    def get_serializer_class(self):
        if self.action == 'list':
            return UserSerializer
        elif self.action == 'retrieve':
            return AthleteDetailSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        qs = self.queryset
        type_user = self.request.query_params.get('type', None)
        qs = qs.annotate(
            runs_finished_count=Count('run', filter=Q(run__status='finished')))
        if type_user == 'coach':
            qs = qs.filter(is_staff=True, is_superuser=False)
        elif type_user == 'athlete':
            qs = qs.filter(is_staff=False, is_superuser=False)
        else:
            qs = qs.filter(is_superuser=False)
        return qs


class StartRunAPIView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status != 'init':
            return Response(status=status.HTTP_400_BAD_REQUEST)
        else:
            run.status = 'in_progress'
            run.save()
            return Response({"status": "updated"})


class StopRunAPIView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status != 'in_progress':
            return Response(status=status.HTTP_400_BAD_REQUEST)
        else:
            positions = run.position_set.all()

            total_distance = 0.0
            prev_point = None

            for position in positions:
                current_point = (position.latitude, position.longitude)
                if prev_point:
                    total_distance += geodesic(prev_point,
                                               current_point).kilometers
                prev_point = current_point

            run.distance = total_distance
            run.status = 'finished'
            run.save()

            finished_runs = Run.objects.filter(athlete=run.athlete,
                                               status='finished').count()
            if finished_runs == 10:
                Challenge.objects.get_or_create(
                    athlete=run.athlete,
                    defaults={
                        'full_name': 'Сделай 10 Забегов!'
                    }
                )

            total_distance_sum = Run.objects.filter(
                athlete=run.athlete,
                status='finished'
            ).aggregate(Sum('distance'))['distance__sum'] or 0

            if total_distance_sum >= 50:
                Challenge.objects.get_or_create(
                    athlete=run.athlete,
                    defaults={
                        'full_name': 'Пробеги 50 километров!'
                    }
                )

            first_position = Position.objects.filter(run=run).order_by(
                'date_time').first()
            last_position = Position.objects.filter(run=run).order_by(
                'date_time').last()

            if first_position and last_position:
                time_difference = last_position.date_time - first_position.date_time
                run.run_time_seconds = time_difference.total_seconds()
            else:
                run.run_time_seconds = 0
            run.save()

            return Response({"status": "updated"})


class AthleteInfoAPIView(APIView):
    def get(self, request, user_id):
        user = get_object_or_404(User, id=user_id)
        athlete_info, created = AthleteInfo.objects.get_or_create(user_id=user)
        serializer = AthleteInfoSerializer(athlete_info)
        return Response(serializer.data)

    def put(self, request, user_id):
        user = get_object_or_404(User, id=user_id)
        weight = request.data.get('weight')
        goals = request.data.get('goals')
        if weight is None or (weight.isdigit() and 0 < int(weight) < 900):
            athlete_info, created = AthleteInfo.objects.update_or_create(
                user_id=user,
                defaults={
                    'weight': weight,
                    'goals': goals,
                }
            )
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        serializer = AthleteInfoSerializer(athlete_info)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ChallengeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['athlete']


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['run']

    def perform_create(self, serializer):
        position = serializer.save()
        athlete_position = (position.latitude, position.longitude)
        for item in CollectibleItem.objects.all():
            item_position = (item.latitude, item.longitude)
            distance_to_item = geodesic(item_position, athlete_position).meters
            if distance_to_item <= 100:
                item.users.add(position.run.athlete_id)


class CollectibleItemViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer


@api_view(['POST'])
def upload_view(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_xlsx_file = request.FILES['file']
        wb = load_workbook(uploaded_xlsx_file, data_only=True)
        sheet = wb.active
        wrong_rows_list = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row,
                                   values_only=True):
            name, uid, value, latitude, longitude, picture = row
            data = {
                'name': name,
                'uid': uid,
                'latitude': latitude,
                'longitude': longitude,
                'picture': picture,
                'value': value,
            }
            serializer = CollectibleItemSerializer(data=data)
            if serializer.is_valid():
                CollectibleItem.objects.create(name=name,
                                               uid=uid,
                                               value=value,
                                               latitude=latitude,
                                               longitude=longitude,
                                               picture=picture)
            else:
                wrong_rows_list.append(
                    [name, uid, value, latitude, longitude, picture])

        return Response(wrong_rows_list)
    return Response([])
